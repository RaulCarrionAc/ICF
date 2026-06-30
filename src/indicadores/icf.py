import pandas as pd
import numpy as np
import holidays
from pathlib import Path

def crear_df_icf(df_a1:pd.DataFrame, df_conteo:pd.DataFrame)->pd.DataFrame:
    df_icf = pd.merge(
        df_a1,
        df_conteo,
        on = ["Fecha","servicio","sentido","tipo_dia","periodo"],
        how="left",
    )

    df_icf = df_icf.rename(columns = {"expediciones_observadas": "EO"})
    df_icf["EO"] = df_icf["EO"].fillna(0)
    df_icf['ICF'] = np.floor(np.minimum(df_icf['EE'], df_icf['EO']) / df_icf['EE'] * 100 + 0.5) / 100
    return df_icf

def calcular_psi(df,fecha_columna="Fecha",fecha_inicio_operacion=None,mas_de_24_meses=None):
    """
    Calcula el parámetro psi (ψ) según la antigüedad de la operación,
    de acuerdo a lo indicado en la Res. 49/2024 (pág. 41):
        - Hasta el mes 24 de operación: psi = 0,90
        - Desde el mes 25 en adelante:  psi = 0,95

    Parámetros
    ----------
    df : DataFrame
        DataFrame que contiene la columna de fecha sobre la que se calculará psi.
    fecha_columna : str
        Nombre de la columna de fecha en df (default 'Fecha').
    fecha_inicio_operacion : str o Timestamp, opcional
        Fecha de inicio de operación del Perímetro de Exclusión. Si se entrega,
        psi se calcula automáticamente fila por fila según la antigüedad real
        a la fecha de cada registro.
    mas_de_24_meses : bool, opcional
        Si no quieres calcular la antigüedad real (por ejemplo, para otros
        servicios donde solo sabes "ya superó los 24 meses" o no), puedes
        forzar el valor directamente: True -> psi=0.95 para todo el df,
        False -> psi=0.90 para todo el df.

    Debes entregar UNO de los dos parámetros: fecha_inicio_operacion o mas_de_24_meses.

    Retorna
    -------
    Series con el valor de psi para cada fila de df.
    """
    if fecha_inicio_operacion is None and mas_de_24_meses is None:
        raise ValueError(
            "Debes especificar 'fecha_inicio_operacion' o 'mas_de_24_meses'."
        )

    if mas_de_24_meses is not None:
        # Modo simple: se fuerza el mismo psi para todo el dataframe
        return pd.Series(0.95 if mas_de_24_meses else 0.90, index=df.index)

    # Modo automático: calcular antigüedad real fila por fila
    fecha_inicio_operacion = pd.Timestamp(fecha_inicio_operacion)

    def mes_operacion(fecha):
        return (fecha.year - fecha_inicio_operacion.year) * 12 + (fecha.month - fecha_inicio_operacion.month) + 1

    meses_operacion = df[fecha_columna].apply(mes_operacion)
    return np.where(meses_operacion <= 24, 0.90, 0.95)

def tabla_periodo_vs_fecha(df_icf, servicio, sentido):
    """
    Reproduce la tabla 'periodo x fecha' (con promedio final) usando la
    razón cruda EO/EE (SIN aplicar el min/cap del ICF normativo) — esto
    es lo que muestra el reporte de referencia, no el ICF oficial.
    """
    df_filtro = df_icf[
        (df_icf['servicio'] == servicio) &
        (df_icf['sentido'] == sentido)
    ].copy()

    # Razón cruda, sin el cap de min(EE,EO)
    df_filtro['ratio_crudo'] = df_filtro['EO'] / df_filtro['EE']

    tabla = df_filtro.pivot_table(
        index='periodo',
        columns='Fecha',
        values='ratio_crudo',
        aggfunc='mean'   # por si hay más de una fila por periodo-fecha (no debería)
    )

    # Formatear columnas de fecha igual al reporte (YYYY-MM-DD)
    tabla.columns = [c.strftime('%Y-%m-%d') for c in tabla.columns]

    # Agregar columna de promedio simple por periodo (a través de todas las fechas)
    tabla['Promedio'] = tabla.mean(axis=1).round(2)

    return tabla

def aplicar_regla_pago(icf, psi=0.95):
    if icf < 0.50:
        return 0.50
    elif icf > psi:
        return 1.00
    else:
        return icf


def construir_resumenes_icf(df_icf: pd.DataFrame, psi_valor: float = 0.95):
    """
    Construye los 4 resúmenes del ICF a nivel mensual:

    1. tabla_por_tipo_demanda: ICF promedio por tipo de demanda
       (df_icf.groupby("tipo_demanda")["ICF"].mean().round(2))
    2. icf_general: promedio simple de los promedios anteriores
       (tabla_por_tipo_demanda.mean().round(3))
    3. tabla_por_tipo_demanda_servicio: ICF promedio por tipo de demanda
       y servicio
       (df_icf.groupby(["tipo_demanda","servicio"])["ICF"].mean().round(2))
    4. icf_pago: promedio simple redondeado a 2 decimales y con regla de pago aplicada

    Retorna (tabla_por_tipo_demanda, icf_general, tabla_por_tipo_demanda_servicio, icf_pago)
    """
    tabla_por_tipo_demanda = df_icf.groupby("tipo_demanda")["ICF"].mean().round(2)
    icf_general = round(tabla_por_tipo_demanda.mean(), 3)
    tabla_por_tipo_demanda_servicio = (
        df_icf.groupby(["tipo_demanda", "servicio"])["ICF"].mean().round(2)
    )
    
    # Promedio de los tipos de demanda redondeado a 2 decimales
    icf_promedio_crudo = round(tabla_por_tipo_demanda.mean(), 2)
    # Aplicar la regla de pago
    icf_pago = aplicar_regla_pago(icf_promedio_crudo, psi=psi_valor)

    return tabla_por_tipo_demanda, icf_general, tabla_por_tipo_demanda_servicio, icf_pago


def exportar_resumenes_icf(
    tabla_por_tipo_demanda: pd.Series,
    icf_general: float,
    tabla_por_tipo_demanda_servicio: pd.Series,
    icf_pago: float,
    ruta_archivo: Path
) -> None:
    """
    Exporta los 4 resúmenes del ICF a un único Excel, cada uno en su
    propia hoja.
    """
    df_tipo_demanda = tabla_por_tipo_demanda.reset_index().rename(
        columns={"ICF": "ICF_promedio"}
    )
    df_general = pd.DataFrame({"ICF_general": [icf_general]})
    df_tipo_demanda_servicio = tabla_por_tipo_demanda_servicio.reset_index().rename(
        columns={"ICF": "ICF_promedio"}
    )
    df_pago = pd.DataFrame({"ICF_pago": [icf_pago]})

    with pd.ExcelWriter(ruta_archivo, engine="openpyxl") as writer:
        df_tipo_demanda.to_excel(writer, sheet_name="Por_TipoDemanda", index=False)
        df_general.to_excel(writer, sheet_name="ICF_General", index=False)
        df_pago.to_excel(writer, sheet_name="ICF_General_Pago", index=False)
        df_tipo_demanda_servicio.to_excel(
            writer, sheet_name="Por_TipoDemanda_Servicio", index=False
        )


def proyectar_simulado_estocastico(
    df_icf_completo: pd.DataFrame,
    df_historico: pd.DataFrame,
    ultima_fecha,
    seed: int = 42
) -> pd.DataFrame:
    """
    Rellena los días faltantes del mes utilizando un muestreo estocástico con reemplazo
    del ratio de cumplimiento (EO / EE) obtenido del histórico acumulado y del mes en curso.
    """
    df_sim = df_icf_completo.copy()
    missing_mask = df_sim["Fecha"] > ultima_fecha

    # Pool de datos observados (historial + mes actual hasta la fecha)
    df_pool_actual = df_sim[~missing_mask].copy()
    if df_historico is not None and not df_historico.empty:
        # Filtrar solo columnas necesarias para evitar problemas de compatibilidad
        columnas_comunes = ["servicio", "sentido", "periodo", "tipo_dia", "tipo_demanda", "EE", "EO"]
        df_hist_filtrado = df_historico[columnas_comunes].copy()
        df_act_filtrado = df_pool_actual[columnas_comunes].copy()
        df_pool = pd.concat([df_hist_filtrado, df_act_filtrado], ignore_index=True)
    else:
        df_pool = df_pool_actual

    if df_pool.empty:
        # Fallback si no hay ningún dato: asumimos cumplimiento perfecto
        df_sim.loc[missing_mask, "EO"] = df_sim.loc[missing_mask, "EE"]
        df_sim.loc[missing_mask, "ICF"] = 1.0
        return df_sim

    # Calcular ratio de cumplimiento observado en el pool
    df_pool["ratio"] = df_pool["EO"] / df_pool["EE"]
    df_pool["ratio"] = df_pool["ratio"].fillna(0.0).clip(lower=0.0)

    # Agrupaciones en cascada para el muestreo
    pool_dict_primary = df_pool.groupby(["servicio", "sentido", "periodo", "tipo_dia"])["ratio"].apply(list).to_dict()
    pool_dict_secondary = df_pool.groupby(["tipo_demanda", "periodo"])["ratio"].apply(list).to_dict()
    pool_dict_tertiary = df_pool.groupby(["periodo"])["ratio"].apply(list).to_dict()

    # Inicializar generador aleatorio para reproducibilidad
    rng = np.random.default_rng(seed)

    df_missing = df_sim[missing_mask].copy()
    simulated_ratios = []

    for idx, row in df_missing.iterrows():
        key1 = (row["servicio"], row["sentido"], row["periodo"], row["tipo_dia"])
        key2 = (row["tipo_demanda"], row["periodo"])
        key3 = row["periodo"]

        ratio_list = pool_dict_primary.get(key1)
        if not ratio_list:
            ratio_list = pool_dict_secondary.get(key2)
        if not ratio_list:
            ratio_list = pool_dict_tertiary.get(key3)

        if ratio_list:
            ratio_val = rng.choice(ratio_list)
        else:
            ratio_val = 1.0  # Fallback absoluto

        simulated_ratios.append(ratio_val)

    # Asignar expediciones observadas simuladas (sin superar las exigidas y redondeado a entero)
    df_sim.loc[missing_mask, "EO"] = np.minimum(
        df_missing["EE"],
        np.round(df_missing["EE"] * simulated_ratios)
    )

    # Recalcular el ICF para los días proyectados
    df_sim.loc[missing_mask, "ICF"] = np.floor(
        np.minimum(df_sim.loc[missing_mask, "EE"], df_sim.loc[missing_mask, "EO"]) 
        / df_sim.loc[missing_mask, "EE"] * 100 + 0.5
    ) / 100

    return df_sim


def proyectar_ideal(df_icf_completo: pd.DataFrame, ultima_fecha) -> pd.DataFrame:
    """
    Rellena los días faltantes del mes asumiendo un escenario ideal en el cual
    se cumple exactamente la frecuencia exigida (ICF = 1.0).
    """
    df_ideal = df_icf_completo.copy()
    missing_mask = df_ideal["Fecha"] > ultima_fecha

    # En el escenario ideal, el cumplimiento de los días faltantes es perfecto
    df_ideal.loc[missing_mask, "EO"] = df_ideal.loc[missing_mask, "EE"]
    df_ideal.loc[missing_mask, "ICF"] = 1.0

    return df_ideal


def crear_tabla_comparativa(
    tabla_demanda_obs: pd.Series, icf_general_obs: float, icf_pago_obs: float,
    tabla_demanda_sim: pd.Series, icf_gen_sim: float, icf_pag_sim: float,
    tabla_demanda_ideal: pd.Series, icf_gen_ideal: float, icf_pag_ideal: float
) -> pd.DataFrame:
    """
    Construye un DataFrame comparativo side-by-side de las métricas clave.
    """
    filas = []

    # 1. Agregar ICF General
    filas.append({
        "Métrica": "ICF General (3 decimales)",
        "Real Observado (a la fecha)": icf_general_obs,
        "Proyección Simulado (Mes Completo)": icf_gen_sim,
        "Proyección Ideal (Mes Completo)": icf_gen_ideal
    })

    # 2. Agregar ICF Pago
    filas.append({
        "Métrica": "ICF Pago (Regla de Pago)",
        "Real Observado (a la fecha)": icf_pago_obs,
        "Proyección Simulado (Mes Completo)": icf_pag_sim,
        "Proyección Ideal (Mes Completo)": icf_pag_ideal
    })

    # 3. Agregar tipos de demanda
    tipos_demanda = sorted(list(
        set(tabla_demanda_obs.index) | 
        set(tabla_demanda_sim.index) | 
        set(tabla_demanda_ideal.index)
    ))

    for td in tipos_demanda:
        val_obs = tabla_demanda_obs.get(td, np.nan)
        val_sim = tabla_demanda_sim.get(td, np.nan)
        val_ideal = tabla_demanda_ideal.get(td, np.nan)

        filas.append({
            "Métrica": f"ICF Promedio Demanda: {td}",
            "Real Observado (a la fecha)": val_obs,
            "Proyección Simulado (Mes Completo)": val_sim,
            "Proyección Ideal (Mes Completo)": val_ideal
        })

    return pd.DataFrame(filas)


def agregar_hoja_simulacion(workbook, df_sim: pd.DataFrame, sheet_name: str, ultima_fecha) -> None:
    """
    Crea una hoja en el libro de Excel y escribe los datos de simulación día a día
    organizados en tablas paralelas (side-by-side) por tipo de demanda, incluyendo
    fórmulas de promedio de Excel, formato condicional de 3 colores y bordes naranjas
    para los días proyectados.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule

    # Crear la hoja
    ws = workbook.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True

    # Orden lógico de las demandas
    demand_order = ["BAJA", "MEDIA", "ALTA"]
    tipos_demanda = [td for td in demand_order if td in df_sim["tipo_demanda"].dropna().unique()]

    # Mapeo de sentidos
    sentido_map = {"I": "Ida", "R": "Reg"}

    # Bordes de distinción
    orange_side = Side(border_style="thin", color="FFA500")
    sim_border = Border(top=orange_side, bottom=orange_side, left=orange_side, right=orange_side)
    
    gray_side = Side(border_style="thin", color="D3D3D3")
    real_border = Border(top=gray_side, bottom=gray_side, left=gray_side, right=gray_side)

    align_center = Alignment(horizontal="center", vertical="center")
    
    start_col = 1
    average_cells = []

    for td in tipos_demanda:
        # Filtrar por tipo de demanda
        df_td = df_sim[df_sim["tipo_demanda"] == td].copy()
        if df_td.empty:
            continue

        # Ordenar cronológicamente
        df_td = df_td.sort_values(["Fecha", "periodo"])

        # Obtener las combinaciones únicas de servicio/sentido para esta demanda
        combinaciones = df_td[["servicio", "sentido"]].drop_duplicates().sort_values(["servicio", "sentido"])
        if combinaciones.empty:
            continue

        # Pivotear para obtener Fecha/Periodo como índice y (servicio, sentido) como columnas
        pivot = df_td.pivot_table(
            index=["Fecha", "periodo"],
            columns=["servicio", "sentido"],
            values="ICF",
            aggfunc="first"
        )
        pivot = pivot.sort_index(level=["Fecha", "periodo"])

        # Escribir cabeceras
        ws.cell(row=2, column=start_col + 1, value="Promedio de Frecuencia").font = Font(bold=True)
        ws.cell(row=2, column=start_col + 3, value="Servicio_dicc").font = Font(bold=True)
        ws.cell(row=2, column=start_col + 4, value="Sentido").font = Font(bold=True)

        ws.cell(row=4, column=start_col, value="Dia").font = Font(bold=True)
        ws.cell(row=4, column=start_col + 1, value="Fecha").font = Font(bold=True)
        ws.cell(row=4, column=start_col + 2, value="Periodo_dicc").font = Font(bold=True)

        ws.cell(row=4, column=start_col).alignment = align_center
        ws.cell(row=4, column=start_col + 1).alignment = align_center
        ws.cell(row=4, column=start_col + 2).alignment = align_center

        data_cols = []
        for idx, (_, row_comb) in enumerate(combinaciones.iterrows()):
            srv = row_comb["servicio"]
            sent = row_comb["sentido"]
            col_idx = start_col + 3 + idx
            data_cols.append(col_idx)

            # Escribir servicio en fila 3
            ws.cell(row=3, column=col_idx, value=srv).font = Font(bold=True)
            ws.cell(row=3, column=col_idx).alignment = align_center
            # Escribir sentido en fila 4
            ws.cell(row=4, column=col_idx, value=sentido_map.get(sent, sent)).font = Font(bold=True)
            ws.cell(row=4, column=col_idx).alignment = align_center

        # Escribir filas de datos
        num_rows = len(pivot)
        for row_idx, ((fecha, per), row_vals) in enumerate(pivot.iterrows()):
            excel_row = 5 + row_idx
            fecha_val = pd.to_datetime(fecha)
            is_simulated = (fecha_val > pd.to_datetime(ultima_fecha))
            current_border = sim_border if is_simulated else real_border

            # Columna Dia (Fórmula TEXT)
            fecha_cell = f"{get_column_letter(start_col + 1)}{excel_row}"
            cell_dia = ws.cell(row=excel_row, column=start_col, value=f'=TEXT({fecha_cell},"dddd")')
            cell_dia.alignment = align_center
            cell_dia.border = current_border

            # Columna Fecha
            cell_fecha = ws.cell(row=excel_row, column=start_col + 1, value=fecha_val.date())
            cell_fecha.number_format = 'dd-mm-yy'
            cell_fecha.alignment = align_center
            cell_fecha.border = current_border

            # Columna Periodo
            cell_per = ws.cell(row=excel_row, column=start_col + 2, value=int(per))
            cell_per.alignment = align_center
            cell_per.border = current_border

            # Columnas de datos (ICF)
            for idx, (_, row_comb) in enumerate(combinaciones.iterrows()):
                srv = row_comb["servicio"]
                sent = row_comb["sentido"]
                val = row_vals.get((srv, sent))
                col_idx = start_col + 3 + idx

                cell_val = ws.cell(row=excel_row, column=col_idx)
                cell_val.alignment = align_center
                cell_val.border = current_border

                if pd.notna(val):
                    cell_val.value = float(val)
                    cell_val.number_format = '0.00'

        # Escribir el promedio de este tipo de demanda
        end_col = start_col + 2 + len(combinaciones)
        avg_col = end_col + 2

        ws.cell(row=3, column=avg_col, value=f'Promedio simulacion "{td}"').font = Font(bold=True)
        ws.cell(row=3, column=avg_col).alignment = align_center

        # Fórmula del promedio de la tabla
        first_data_cell = f"{get_column_letter(start_col + 3)}5"
        last_data_cell = f"{get_column_letter(end_col)}{5 + num_rows - 1}"
        ws.cell(row=4, column=avg_col, value=f'=AVERAGE({first_data_cell}:{last_data_cell})').font = Font(bold=True)
        ws.cell(row=4, column=avg_col).alignment = align_center
        ws.cell(row=4, column=avg_col).number_format = '0.00000'

        average_cells.append(f"{get_column_letter(avg_col)}4")

        # Formato condicional (escala de 3 colores) para el rango de datos
        cell_range = f"{get_column_letter(start_col + 3)}5:{get_column_letter(end_col)}{5 + num_rows - 1}"
        color_scale = ColorScaleRule(
            start_type='min', start_color='FFF8696B', # Rojo suave
            mid_type='percentile', mid_value=50, mid_color='FFFFEB84', # Amarillo suave
            end_type='max', end_color='FF63BE7B' # Verde suave
        )
        ws.conditional_formatting.add(cell_range, color_scale)

        # Siguiente tabla empieza 5 columnas después
        start_col = end_col + 5

    # Escribir resumen general al final
    if average_cells:
        summary_label_col = start_col - 2
        summary_val_col = start_col - 1

        ws.cell(row=3, column=summary_label_col, value="ICF simulado").font = Font(bold=True)
        ws.cell(row=3, column=summary_label_col).alignment = align_center

        avg_formula_terms = "+".join(average_cells)
        formula = f"=({avg_formula_terms})/{len(average_cells)}"
        cell_summary = ws.cell(row=3, column=summary_val_col, value=formula)
        cell_summary.font = Font(bold=True)
        cell_summary.alignment = align_center
        cell_summary.number_format = '0.00000'


def exportar_reporte_proyeccion(
    df_comparativa: pd.DataFrame,
    df_sim: pd.DataFrame,
    df_ideal: pd.DataFrame,
    tabla_demanda_sim: pd.Series, icf_gen_sim: float, tabla_serv_sim: pd.Series, icf_pag_sim: float,
    tabla_demanda_ideal: pd.Series, icf_gen_ideal: float, tabla_serv_ideal: pd.Series, icf_pag_ideal: float,
    ultima_fecha,
    ruta_archivo: Path
) -> None:
    """
    Exporta todos los resultados de simulación y proyecciones a un archivo Excel.
    """
    df_demanda_sim = tabla_demanda_sim.reset_index().rename(columns={"ICF": "ICF_promedio"})
    df_serv_sim = tabla_serv_sim.reset_index().rename(columns={"ICF": "ICF_promedio"})
    df_gen_sim = pd.DataFrame({"ICF_general": [icf_gen_sim], "ICF_pago": [icf_pag_sim]})

    df_demanda_ideal = tabla_demanda_ideal.reset_index().rename(columns={"ICF": "ICF_promedio"})
    df_serv_ideal = tabla_serv_ideal.reset_index().rename(columns={"ICF": "ICF_promedio"})
    df_gen_ideal = pd.DataFrame({"ICF_general": [icf_gen_ideal], "ICF_pago": [icf_pag_ideal]})

    with pd.ExcelWriter(ruta_archivo, engine="openpyxl") as writer:
        df_comparativa.to_excel(writer, sheet_name="Comparativa", index=False)
        df_gen_sim.to_excel(writer, sheet_name="Simulado_ICF_General", index=False)
        df_demanda_sim.to_excel(writer, sheet_name="Simulado_Por_TipoDemanda", index=False)
        df_serv_sim.to_excel(writer, sheet_name="Simulado_Por_Servicio", index=False)
        df_gen_ideal.to_excel(writer, sheet_name="Ideal_ICF_General", index=False)
        df_demanda_ideal.to_excel(writer, sheet_name="Ideal_Por_TipoDemanda", index=False)
        df_serv_ideal.to_excel(writer, sheet_name="Ideal_Por_Servicio", index=False)

        workbook = writer.book
        agregar_hoja_simulacion(workbook, df_sim, "Simulacion", ultima_fecha)
        agregar_hoja_simulacion(workbook, df_ideal, "Simulacion_Ideal", ultima_fecha)